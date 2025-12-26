from flask import Blueprint, request, jsonify, send_file, current_app
import pandas as pd
import openpyxl
from io import BytesIO
import os
from datetime import datetime
from src.models.user import db, User
from src.models.document import Document, Project, Schedule

excel_bp = Blueprint('excel', __name__)

@excel_bp.route('/export/documents', methods=['GET'])
def export_documents():
    """도면 목록을 엑셀로 내보내기"""
    try:
        project_id = request.args.get('project_id')
        status = request.args.get('status')
        
        query = Document.query
        if project_id:
            query = query.filter(Document.project_id == project_id)
        if status:
            query = query.filter(Document.status == status)
        
        documents = query.all()
        
        # 데이터 준비
        data = []
        for doc in documents:
            data.append({
                'ID': doc.id,
                '제목': doc.title,
                '설명': doc.description or '',
                '파일명': doc.filename,
                '파일 크기 (KB)': round(doc.file_size / 1024, 2) if doc.file_size else 0,
                '파일 형식': doc.file_type,
                '프로젝트 ID': doc.project_id or '',
                '버전': doc.version,
                '상태': doc.status,
                '생성자': doc.creator.username if doc.creator else '',
                '생성일': doc.created_at.strftime('%Y-%m-%d %H:%M:%S') if doc.created_at else '',
                '수정일': doc.updated_at.strftime('%Y-%m-%d %H:%M:%S') if doc.updated_at else ''
            })
        
        # DataFrame 생성
        df = pd.DataFrame(data)
        
        # 엑셀 파일 생성
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='도면 목록', index=False)
            
            # 워크시트 스타일링
            workbook = writer.book
            worksheet = writer.sheets['도면 목록']
            
            # 헤더 스타일
            header_font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            header_fill = openpyxl.styles.PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            # 열 너비 자동 조정
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        filename = f"documents_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@excel_bp.route('/export/projects', methods=['GET'])
def export_projects():
    """프로젝트 목록을 엑셀로 내보내기"""
    try:
        projects = Project.query.all()
        
        # 데이터 준비
        data = []
        for project in projects:
            data.append({
                'ID': project.id,
                '프로젝트명': project.name,
                '설명': project.description or '',
                '선박 유형': project.ship_type or '',
                '고객사': project.client or '',
                '시작일': project.start_date.strftime('%Y-%m-%d') if project.start_date else '',
                '종료일': project.end_date.strftime('%Y-%m-%d') if project.end_date else '',
                '상태': project.status,
                '생성자': project.creator.username if project.creator else '',
                '생성일': project.created_at.strftime('%Y-%m-%d %H:%M:%S') if project.created_at else ''
            })
        
        # DataFrame 생성
        df = pd.DataFrame(data)
        
        # 엑셀 파일 생성
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='프로젝트 목록', index=False)
            
            # 워크시트 스타일링
            workbook = writer.book
            worksheet = writer.sheets['프로젝트 목록']
            
            # 헤더 스타일
            header_font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            header_fill = openpyxl.styles.PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            # 열 너비 자동 조정
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        filename = f"projects_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@excel_bp.route('/export/users', methods=['GET'])
def export_users():
    """사용자 목록을 엑셀로 내보내기"""
    try:
        users = User.query.filter(User.is_active == True).all()
        
        # 데이터 준비
        data = []
        for user in users:
            data.append({
                'ID': user.id,
                '사용자명': user.username,
                '이메일': user.email,
                '성명': user.full_name or '',
                '부서': user.department or '',
                '직책': user.position or '',
                '전화번호': user.phone or '',
                '역할': user.role,
                '언어': user.language,
                '마지막 로그인': user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else '',
                '생성일': user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else ''
            })
        
        # DataFrame 생성
        df = pd.DataFrame(data)
        
        # 엑셀 파일 생성
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='사용자 목록', index=False)
            
            # 워크시트 스타일링
            workbook = writer.book
            worksheet = writer.sheets['사용자 목록']
            
            # 헤더 스타일
            header_font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            header_fill = openpyxl.styles.PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            # 열 너비 자동 조정
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        filename = f"users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@excel_bp.route('/import/projects', methods=['POST'])
def import_projects():
    """엑셀 파일에서 프로젝트 데이터 가져오기"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': '파일이 선택되지 않았습니다.'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': '파일이 선택되지 않았습니다.'}), 400
        
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'error': '엑셀 파일만 업로드 가능합니다.'}), 400
        
        # 엑셀 파일 읽기
        df = pd.read_excel(file)
        
        # 필수 컬럼 확인
        required_columns = ['프로젝트명']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({
                'success': False, 
                'error': f'필수 컬럼이 누락되었습니다: {", ".join(missing_columns)}'
            }), 400
        
        # 데이터 처리
        success_count = 0
        error_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                # 프로젝트 ID 생성 또는 사용
                project_id = row.get('ID', f"PRJ_{datetime.now().strftime('%Y%m%d')}_{index+1:03d}")
                
                # 기존 프로젝트 확인
                existing_project = Project.query.filter_by(id=project_id).first()
                if existing_project:
                    errors.append(f"행 {index+2}: 프로젝트 ID '{project_id}'가 이미 존재합니다.")
                    error_count += 1
                    continue
                
                project = Project(
                    id=project_id,
                    name=row['프로젝트명'],
                    description=row.get('설명', ''),
                    ship_type=row.get('선박 유형', ''),
                    client=row.get('고객사', ''),
                    start_date=pd.to_datetime(row.get('시작일')).date() if pd.notna(row.get('시작일')) else None,
                    end_date=pd.to_datetime(row.get('종료일')).date() if pd.notna(row.get('종료일')) else None,
                    status=row.get('상태', 'active'),
                    created_by=1  # TODO: 실제 사용자 ID로 변경
                )
                
                db.session.add(project)
                success_count += 1
                
            except Exception as e:
                errors.append(f"행 {index+2}: {str(e)}")
                error_count += 1
        
        if success_count > 0:
            db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'{success_count}개 프로젝트가 성공적으로 가져와졌습니다.',
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@excel_bp.route('/template/projects', methods=['GET'])
def download_project_template():
    """프로젝트 가져오기용 엑셀 템플릿 다운로드"""
    try:
        # 템플릿 데이터
        template_data = {
            'ID': ['PRJ_001', 'PRJ_002'],
            '프로젝트명': ['컨테이너선 A호', '벌크선 B호'],
            '설명': ['20,000TEU 컨테이너선', '80,000DWT 벌크선'],
            '선박 유형': ['컨테이너선', '벌크선'],
            '고객사': ['현대상선', '팬오션'],
            '시작일': ['2024-01-01', '2024-02-01'],
            '종료일': ['2024-12-31', '2025-01-31'],
            '상태': ['active', 'active']
        }
        
        df = pd.DataFrame(template_data)
        
        # 엑셀 파일 생성
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='프로젝트 템플릿', index=False)
            
            # 워크시트 스타일링
            workbook = writer.book
            worksheet = writer.sheets['프로젝트 템플릿']
            
            # 헤더 스타일
            header_font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            header_fill = openpyxl.styles.PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
            
            # 열 너비 자동 조정
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='project_template.xlsx'
        )
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

