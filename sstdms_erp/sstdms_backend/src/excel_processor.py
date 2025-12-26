#!/usr/bin/env python3
"""
Excel 파일 처리 유틸리티
도면 리스트 및 프로젝트 데이터 처리
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os
from datetime import datetime, timedelta
import json

class ExcelProcessor:
    def __init__(self):
        self.template_columns = [
            'NO', 'COMMON', 'DWG NAME', 'TYPE', 'START', 'MONTH', 'FINISH', 
            'PROGRESS', 'REV', 'STATUS', 'ISSUED_DATE', 'APPROVED_BY', 'REMARKS'
        ]
    
    def read_drawing_list(self, file_path):
        """도면 리스트 엑셀 파일 읽기"""
        try:
            # 엑셀 파일 읽기
            df = pd.read_excel(file_path, sheet_name=0)
            
            # 컬럼명 정리
            df.columns = df.columns.str.strip()
            
            # 데이터 정리
            drawing_list = []
            for index, row in df.iterrows():
                if pd.isna(row.iloc[0]) or row.iloc[0] == '':
                    continue
                    
                drawing_data = {
                    'no': str(row.iloc[0]) if not pd.isna(row.iloc[0]) else '',
                    'common': str(row.iloc[1]) if len(row) > 1 and not pd.isna(row.iloc[1]) else '',
                    'dwg_name': str(row.iloc[2]) if len(row) > 2 and not pd.isna(row.iloc[2]) else '',
                    'type': str(row.iloc[3]) if len(row) > 3 and not pd.isna(row.iloc[3]) else '',
                    'start_date': row.iloc[4] if len(row) > 4 and not pd.isna(row.iloc[4]) else None,
                    'duration_months': row.iloc[5] if len(row) > 5 and not pd.isna(row.iloc[5]) else 1,
                    'finish_date': row.iloc[6] if len(row) > 6 and not pd.isna(row.iloc[6]) else None,
                    'progress': row.iloc[7] if len(row) > 7 and not pd.isna(row.iloc[7]) else 0,
                    'revision': str(row.iloc[8]) if len(row) > 8 and not pd.isna(row.iloc[8]) else 'A',
                    'status': str(row.iloc[9]) if len(row) > 9 and not pd.isna(row.iloc[9]) else 'DRAFT',
                    'issued_date': row.iloc[10] if len(row) > 10 and not pd.isna(row.iloc[10]) else None,
                    'approved_by': str(row.iloc[11]) if len(row) > 11 and not pd.isna(row.iloc[11]) else '',
                    'remarks': str(row.iloc[12]) if len(row) > 12 and not pd.isna(row.iloc[12]) else ''
                }
                drawing_list.append(drawing_data)
            
            return drawing_list
            
        except Exception as e:
            raise Exception(f"엑셀 파일 읽기 실패: {str(e)}")
    
    def create_drawing_list_template(self, project_id, output_path):
        """도면 리스트 템플릿 생성"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{project_id}_Drawing_List"
            
            # 헤더 스타일
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # 헤더 작성
            for col, header in enumerate(self.template_columns, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = header_border
                cell.alignment = header_alignment
            
            # 샘플 데이터 추가
            sample_data = [
                ['1', 'GENERAL', 'GENERAL ARRANGEMENT', 'BASIC', '2024-01-01', '2', '2024-03-01', '50%', 'A', 'DRAFT', '2024-01-15', 'Kim', 'Initial design'],
                ['2', 'HULL', 'MIDSHIP SECTION', 'BASIC', '2024-01-15', '1.5', '2024-03-01', '30%', 'A', 'DRAFT', '', 'Lee', 'Under review'],
                ['3', 'MACHINERY', 'ENGINE ROOM LAYOUT', 'APPROVAL', '2024-02-01', '3', '2024-05-01', '10%', 'A', 'DRAFT', '', 'Park', 'Pending approval']
            ]
            
            for row_idx, row_data in enumerate(sample_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 컬럼 너비 조정
            column_widths = [5, 12, 25, 12, 12, 8, 12, 10, 5, 10, 12, 12, 20]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            wb.save(output_path)
            return True
            
        except Exception as e:
            raise Exception(f"템플릿 생성 실패: {str(e)}")
    
    def export_drawing_list(self, drawing_data, project_info, output_path):
        """도면 리스트를 엑셀로 내보내기"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{project_info.get('id', 'PROJECT')}_Drawing_List"
            
            # 프로젝트 정보 헤더
            ws.merge_cells('A1:M1')
            title_cell = ws['A1']
            title_cell.value = f"DRAWING LIST - {project_info.get('name', 'PROJECT')}"
            title_cell.font = Font(bold=True, size=14)
            title_cell.alignment = Alignment(horizontal='center')
            
            ws.merge_cells('A2:M2')
            info_cell = ws['A2']
            info_cell.value = f"Client: {project_info.get('client', '')} | Ship Type: {project_info.get('ship_type', '')} | Date: {datetime.now().strftime('%Y-%m-%d')}"
            info_cell.alignment = Alignment(horizontal='center')
            
            # 헤더 스타일
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 헤더 작성 (4행부터)
            for col, header in enumerate(self.template_columns, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = header_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 데이터 작성
            for row_idx, drawing in enumerate(drawing_data, 5):
                row_data = [
                    drawing.get('no', ''),
                    drawing.get('common', ''),
                    drawing.get('dwg_name', ''),
                    drawing.get('type', ''),
                    drawing.get('start_date', ''),
                    drawing.get('duration_months', ''),
                    drawing.get('finish_date', ''),
                    drawing.get('progress', ''),
                    drawing.get('revision', ''),
                    drawing.get('status', ''),
                    drawing.get('issued_date', ''),
                    drawing.get('approved_by', ''),
                    drawing.get('remarks', '')
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            
            # 컬럼 너비 조정
            column_widths = [5, 12, 25, 12, 12, 8, 12, 10, 5, 10, 12, 12, 20]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
            
            wb.save(output_path)
            return True
            
        except Exception as e:
            raise Exception(f"엑셀 내보내기 실패: {str(e)}")
    
    def create_gantt_data(self, drawing_data, project_start_date=None):
        """간트차트용 데이터 생성"""
        try:
            if not project_start_date:
                project_start_date = datetime.now()
            elif isinstance(project_start_date, str):
                project_start_date = datetime.strptime(project_start_date, '%Y-%m-%d')
            
            gantt_data = []
            for drawing in drawing_data:
                # 시작일 계산
                if drawing.get('start_date'):
                    if isinstance(drawing['start_date'], str):
                        start_date = datetime.strptime(drawing['start_date'], '%Y-%m-%d')
                    else:
                        start_date = drawing['start_date']
                else:
                    start_date = project_start_date
                
                # 기간 계산
                duration_months = float(drawing.get('duration_months', 1))
                duration_days = int(duration_months * 30)  # 대략적인 계산
                
                # 종료일 계산
                finish_date = start_date + timedelta(days=duration_days)
                
                gantt_item = {
                    'common': drawing.get('common', ''),
                    'dwg_name': drawing.get('dwg_name', ''),
                    'type': drawing.get('type', 'BASIC'),
                    'start': start_date.strftime('%Y-%m-%d'),
                    'duration_months': duration_months,
                    'finish': finish_date.strftime('%Y-%m-%d'),
                    'progress': float(drawing.get('progress', 0)) if isinstance(drawing.get('progress'), (int, float)) else 0
                }
                gantt_data.append(gantt_item)
            
            return gantt_data
            
        except Exception as e:
            raise Exception(f"간트차트 데이터 생성 실패: {str(e)}")

def process_uploaded_excel(file_path):
    """업로드된 엑셀 파일 처리"""
    processor = ExcelProcessor()
    return processor.read_drawing_list(file_path)

